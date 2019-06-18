from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Color, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from subprocess import check_output
from tkcolorpicker import askcolor
import tkinter as tk
import tkinter.ttk as ttk
import string
import re
import os
home = os.path.expanduser('~')
doc = home + '/Documents/testBuildSheet.xlsx'
wb = Workbook()
sheet = wb.active

#cell formatting
lsideBorders = Border(left=Side(style='thick'))
rsideBorders = Border(right=Side(style='thick'))
bBorders = Border(bottom=Side(style='thick'))
topBorders = Border(top=Side(style='thick'))
tlBorder = Border(top=Side(style='thick'), left=Side(style='thick'))
trBorder = Border(top=Side(style='thick'), right=Side(style='thick'))
blBorder = Border(bottom=Side(style='thick'), left=Side(style='thick'))
brBorder = Border(bottom=Side(style='thick'), right=Side(style='thick'))
iipBorderl = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
iipBorderlt = Border(top=Side(style='thick'), bottom=Side(style='thin'), left=Side(style='thin'))
iipBorderrt = Border(top=Side(style='thick'), bottom=Side(style='thin'), right=Side(style='thin'))
iipBorderr = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
backgroundColor = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type='solid')
titleFill = PatternFill(start_color='ccffff', end_color='ccffff', fill_type='solid')
titleFont = Font(name='Calibri', size=14, bold=True)

cwidths = {'A': 2.57, 'B': 23, 'C': 20, 'D': 15.57, 'E': 17.71, 'F': 16.29, 'G': 13, 'H': 10.57, 'I': 12.43, 'J': 10, 'K': 19.71, 'L': 8.57}
headings = ['System', 'Type', 'Name', 'Interface', 'IP', 'User ID', 'PWD', 'Console', 'State', 'Switch::port#', 'Rack']
yesList = ['YES', 'Y']
noList = ['NO', 'N']
dupList = ['DUP', 'DUPLICATE', 'COPY', 'SAME']
fixList = ['FIX', 'CHANGE']
mainList = list()
portList = list()
newLocation = 5
name = ''
backupName=''
loc = ''
#instruction for duplicate
dup = list()


def setBackgroundColor(c):
    try:
        global backgroundColor
        backgroundColor = PatternFill(start_color=c, end_color=c, fill_type='solid')
    except:
        print("Wrong color")


"""
addDeviceLayout()
Arguments:
    rs: Row start. At what row the device layout begins
    rend: Row end. At what row th device layout ends
Description:
Formats the cells in excel to create the device entry.

"""


def addDeviceLayout(rs, rend):
    t = 0
    for y in range(2, 13):
        for x in range(rs, rend+1):
            sheet.cell(row=x, column=y).fill = backgroundColor
            #first row
            if x == rs:
                # if y == 4:
                #     sheet.cell(row=x, column=y).value=name
                if y == 2:
                    sheet.cell(row=x, column=y).value=name
                    sheet.cell(row=x, column=y).border = tlBorder
                elif y == 11 or y == 12:
                    sheet.cell(row=x, column=y).border = trBorder
                else:
                    sheet.cell(row=x, column=y).border = topBorders
            #last row
            elif x == rend:
                if y == 2:
                    sheet.cell(row=x, column=y).border = blBorder
                elif y == 11 or y == 12:
                    sheet.cell(row=x, column=y).border = brBorder
                else:
                    sheet.cell(row=x, column=y).border = bBorders
            #all other rows
            else:
                if y == 2:
                    sheet.cell(row=x, column=y).border = lsideBorders
                elif y == 11 or y == 12:
                        sheet.cell(row=x, column=y).border = rsideBorders
        #area for ports
        for x in range(rs, rend):  
            if y == 5:
                if x == rs:
                    sheet.cell(row=x, column=y).border = iipBorderlt
                else:
                    sheet.cell(row=x, column=y).border = iipBorderl
                if len(portList) >= t:
                    sheet.cell(row=x, column=y).value = portList[t]
                t += 1
            elif y == 6:
                if x == rs:
                    sheet.cell(row=x, column=y).border = iipBorderrt
                else:
                    sheet.cell(row=x, column=y).border = iipBorderr
    sheet.merge_cells(start_row=rs, start_column=12, end_row=rend, end_column=12)
    sheet.cell(row=rs, column=12).alignment= Alignment(horizontal='center', vertical='center')
    sheet.cell(row=rs, column=12).value = loc    


"""
makeHeader()
Arguments:
    cs: Column Start (usually 4 for the header)
    cend: Column End (usually 13 for the header)
    r: Row number that the header will be placed on
Description:
Formats the cells in excel to create a header

"""


def makeHeader(r, cs=2, cend=12):
    t = 0
    for y in range(cs, cend+1):
        sheet.cell(row=r, column=y).fill = titleFill
        sheet.cell(row=r, column=y).border = topBorders
        if y == cs:
            sheet.cell(row=r, column=y).border = tlBorder
        elif y == cend-1 or y == cend:
            sheet.cell(row=r, column=y).border = trBorder
        if t < len(headings):
            sheet.cell(row=r, column=y).value = headings[t]
        t += 1
        sheet.cell(row=r, column=y).font = titleFont

"""
chooseColor()
Arguments: none
Description:
    Uses tkinter color selection form to allow user to select color they want the device information
    to appear in.
"""


def chooseColor():
    root = tk.Tk()
    style = ttk.Style(root)
    style.theme_use('clam')
    root.lift()
    root.attributes('-topmost', True)
    root.focus_force()
    
    color = askcolor((243,205,121), root)[1]
    if color != '' and color != None:
        color = color[1:]
        setBackgroundColor(color)
        # global backgroundColor
        # backgroundColor = PatternFill(start_color=color, end_color=color, fill_type='solid')
    root.destroy()
    root.mainloop()


"""
getPorts()
Arguments:None
Description:
    Asks user for list of ports and parses returned information.
"""


def getPorts():
    global portList
    ports = input("Ports on device (ex. e0a, e0b | e0a-e0z | Port 1-Port 10): ")
    if ',' in ports:
        p = ports.split(',')
    else:
        p = {ports}

    for i in p:
        x = re.search(r"[e][0-9][a-z][-][e][0-9][a-z]", i)
        if '-' not in i:
            portList.append(i)
        #test if port id starts in 'e'
        #elif i[0] == 'e'  or i[1] == 'e':
        elif x != None:
            beginningList = list()
            endList = list()
            b=0
            e=0
            x = re.search(r"[e][0-9][a-z][-][e][0-9][a-z]", i)
            #if x != None:
            entry = x.string.split('-')
            beginningList.append(entry[0][-3] + entry[0][-2])
            beginningList.append(entry[1][-3] + entry[1][-2])
            endList.append(entry[0][-1])
            endList.append(entry[1][-1])
            if beginningList[0] == beginningList[1]:
                start = 0
                stop = 0
                for l in string.ascii_lowercase:
                    if l == endList[0]:
                        start = string.ascii_lowercase.index(l)
                    elif l == endList[1]:
                        stop = string.ascii_lowercase.index(l)
                if start == stop:
                    portList.append(beginningList[0] + endList[0])
                elif start > stop:
                    e = start
                    b = stop
                elif start < stop:
                    b = start
                    e = stop
                for x in range(b+1, e):
                    endList.append(string.ascii_lowercase[x])
                endList.sort()
                for e in endList:
                    portList.append(beginningList[0] + e)
        #test if port id ends in digit
        elif i[-1] in string.digits:
            letterList = list()
            endpoints = list()
            counter = 0
            beginning = 0
            end = 0
                
            x = re.search(r"[0-9]+[a-zA-Z]*[ ]*[a-zA-Z]*[ ]+[0-9]+", i)
            if x != None:
                findRange = x.string.split('-')
                newRange = findRange[0].split(' ')
                secondRange = findRange[1].split(' ')
                if newRange[0] != '':
                    if str.isalpha(newRange[1]):
                        letterList.append(newRange[0] + newRange[1])
                        endpoints.append(int(newRange[2]))
                    else:
                        letterList.append(newRange[0])
                        endpoints.append(int(newRange[1]))
                else: 
                    if str.isalpha(newRange[2]):
                        letterList.append(newRange[1] + " " + newRange[2])
                        endpoints.append(int(newRange[3]))
                    else:
                        letterList.append(newRange[1])
                        endpoints.append(int(newRange[2]))
                if len(secondRange) > 1:
                    if str.isalpha(secondRange[1]):
                        letterList.append(secondRange[0] + " " + secondRange[1])
                        endpoints.append(int(secondRange[2]))
                    else:
                        letterList.append(secondRange[0])
                        endpoints.append(int(secondRange[1]))
                else:
                    letterList.append(newRange[0])
                    endpoints.append(int(secondRange[0]))
            # if endpoints != None:
            #     if endpoints[1] > endpoints[0]:
            #             end = endpoints[1]
            #             beginning = endpoints[0]
            #     elif endpoints[0] > endpoints[1]:
            #             beginning = endpoints[1]
            #             end = endpoints[0]
            #     counter = end - beginning
            # #adds the first entry to the list [letter][number]
            # if counter == 0 and (letterList is not None and endpoints is not None):
            #     portList.append(letterList[0] + endpoints[0])
            # elif counter == 0 and letterList is None and endpoints is not None:
            #     portList.append(endpoints[0])
            # #cycles through to add the rest of the entries
            # elif counter > 0:
            #     for x in range(beginning, end + 1):
            #         portList.append(letterList[0] + " {}".format(x))
        



            x = re.search(r"[a-zA-Z]*[ ]*[0-9]+", i)
            if x != None:
                findRange = x.string.split('-')
                #gets first letter from both sides of '-'
                for r in findRange:
                        letterList.append(str(''.join(list(filter(str.isalpha, r)))))
                #checks if there is a letter to the right of '-', and if not simply takes the number
                if letterList[1] is not None and letterList[1] != "":
                        if letterList[0].upper() == letterList[1].upper():
                                for r in findRange:
                                        endpoints.append(int(''.join(list(filter(str.isdigit, r)))))
                elif letterList[1] is None or letterList[1] == "":
                        for r in findRange:
                                endpoints.append(int(''.join(list(filter(str.isdigit, r)))))
                #compares the 2 numbers so that they can be traversed even if in the wrong order
                if endpoints != None:
                        if endpoints[1] > endpoints[0]:
                                end = endpoints[1]
                                beginning = endpoints[0]
                        elif endpoints[0] > endpoints[1]:
                                beginning = endpoints[1]
                                end = endpoints[0]
                        counter = end - beginning
                #adds the first entry to the list [letter][number]
                if counter == 0 and (letterList is not None and endpoints is not None):
                        portList.append(letterList[0] + endpoints[0])
                elif counter == 0 and letterList is None and endpoints is not None:
                    portList.append(endpoints[0])
                #cycles through to add the rest of the entries
                elif counter > 0:
                        for x in range(beginning, end + 1):
                            y = re.search(r"[a-zA-Z][ ][0-9]+", i)
                            if letterList is not None:
                                if y != None:
                                    portList.append(letterList[0] + " {}".format(x))
                                else:
                                    portList.append(letterList[0] + "{}".format(x))
                            else:
                                portList.append(x)

def clearScreen():
    if os.name == 'nt':
        _ = os.system('cls')
    else:
        _ = os.system('clear')

def printDevices():
    #clearScreen()
    print("These entries will be written:\n")
    x = 1
    for item in mainList:
        print("{0}.Name: {1}\nLocation: {2}\nPortList: {3}\n".format(x, item[0], item[1], str(item[5])))
        x += 1

def writingSpreadsheet():
    printDevices()
    t = False
    while t == False:
        cont = input("\nDoes this look correct?(Y/N) ").upper()
        if cont in noList:
            return False
        elif cont in yesList:
            t = True
        else:
            print("Please type Y or N")
    global name
    global loc
    global backgroundColor
    global portList
    for item in mainList:
        name = item[0]
        loc = item[1]
        backgroundColor = item[4]
        portList = item[5]
        addDeviceLayout(item[2], item[3])
    return True



def fixEntry():
    z = False
    while z == False:
        printDevices()
        listText = "1"
        if len(mainList) > 1:
            listText = "1-{}".format(len(mainList))
        try:
            try:
                f = int(input("Which device entry would you like to fix?\nSelect Number ({}), 0 to quit): ".format(listText)))
            except ValueError:
                print("Please the device number, or 0 to quit.")
                break
            if f > len(mainList) or f < 0:
                print("Device Number not available. Please select from list. ({}), 0 to quit".format(listText))
            elif f == 0:
                z = True
            elif f < len(mainList) and f > 0:
                currentDevice = mainList[f-1]
                cl = False
                while cl == False:
                    print("1. Name ({0})\n2. Location ({1})\n3. Color\n4. Ports ({2})\n".format(currentDevice[0], currentDevice[1], currentDevice[5]))
                    comp = input("Select component to change(q to quit): ").upper()
                    if comp == 'Q':
                        cl = True

        except:
            print("Please enter the number of the device. ({}), 0 to quit".format(listText))



r = True
while r == True:
    fname = input("What would you like the filename to be? ")
    if fname != "":
        doc= home + '/Documents/{}.xlsx'.format(fname)
        if os.path.isfile(doc):
            l = True
            while l == True:
                a = input("File exists. Overwrite?(Y/N): ").upper()
                if a in yesList:
                    try:
                        os.remove(doc)
                        l = False
                        r = False
                    except Exception as e:
                        print("Unable to replace file\n{}".format(str(e)))
                        l = False
                elif a in noList:
                    l = False
                else:
                    print("Incorrect input (Y/N)\n\n")
        else:
            r = False
    else:
        print("Name cannot be blank\n\n")


for c, s in cwidths.items():
    sheet.column_dimensions[c].width = s

makeHeader(4)

d = False
l = True
runtimes = 0
while l == True:
    #create list containing device information
    #[name, loc, startr, endr, color, ports]
    #[name, loc, newlocation, newlocation(edited), backgroundcolor, portlist]
    thisDevice = list()
    if d == False:
        name = input("Type of device: ").upper()
    if runtimes > 0:
        r = True
        if d == False:
            while r == True:
                keepPorts = input("Same ports as last time? ").upper()
                if keepPorts in noList:
                    portList.clear()
                    getPorts()
                    r = False
                elif keepPorts in yesList:
                    print("Keeping same ports: " + str(portList))
                    r = False
                else:
                    print("Incorrect response(Y/N needed)\n")
        else:
            print("Keeping same ports: " + str(portList))
    else:
        getPorts()
    if d == False:
        loc = input("Device rack location: ").upper()
        cloop = True
        while cloop == True:
            col = input("Choose color for device? ").upper()
            if col in yesList:
                chooseColor()
                cloop = False
            if col in noList:
                cloop = False
    #addDeviceLayout(newLocation, len(portList) + newLocation)
    thisDevice.append(name)
    thisDevice.append(loc)
    thisDevice.append(newLocation)
    newLocation += len(portList)
    thisDevice.append(newLocation)
    thisDevice.append(backgroundColor)
    thisDevice.append(portList[:])
    mainList.append(thisDevice)
    newLocation += 1
    loop = True
    while loop == True:
        begin = input("Add another device?(Y/N/Copy/Fix) ").upper()
        if begin in noList:
            if writingSpreadsheet() == False:
                print("Something isn't right")
            else:
                l = False
                loop = False
        elif begin in yesList:
            d = False
            runtimes += 1
            loop = False
        elif begin in dupList:
            d = True
            runtimes += 1
            loop = False
        elif begin in fixList:
            fixEntry()
        else:
            print("Incorrect response(Y/N/Copy needed)\n")




try:
    wb.save(doc)
except Exception as e:
    print("Unable to save Document\n{}".format(str(e)))
check_output(doc, shell=True)